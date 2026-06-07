#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
提示卡分表工具 - 将总表每行拆分为独立 sheet
"""

import sys, os, re, traceback

# PyInstaller hook marker
_hook_openpyxl = ("openpyxl", "openpyxl.styles", "openpyxl.utils",
                  "openpyxl.workbook", "openpyxl.worksheet")


def split_prompt_card(input_path: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"文件不存在：{input_path}")
    ext = os.path.splitext(input_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"不支持的文件格式：{ext}，请使用 .xlsx 文件")

    wb = openpyxl.load_workbook(input_path)
    ws_src = wb.active

    # 表头（第2行）
    headers = [ws_src.cell(row=2, column=c).value for c in range(1, ws_src.max_column + 1)]

    # 收集数据行
    rows_data = []
    for r in range(3, ws_src.max_row + 1):
        if not ws_src.cell(row=r, column=3).value:
            continue
        rows_data.append([ws_src.cell(row=r, column=c).value for c in range(1, ws_src.max_column + 1)])

    if not rows_data:
        raise ValueError("未找到任何有效数据")

    # 配色
    DARK = "1B3A5C"
    ACCENT = "2B65A8"
    BG_TITLE = "E8F0FE"
    VALUE_C = "333333"

    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0'),
    )
    no_top = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style=None),
        bottom=Side(style='thin', color='C0C0C0'),
    )

    def fmt_val(v):
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v).strip() if v is not None else ''

    def calc_lines(text, chars_per_line):
        if not text:
            return 1
        total = 0
        for p in text.split('\n'):
            if not p:
                total += 1
                continue
            w = sum(2 if ord(ch) > 127 else 1 for ch in p)
            total += max(1, -(-w // chars_per_line))
        return total + max(1, int(total * 0.2))  # +20% margin

    def make_label_cell(ws, r, c, label):
        cl = ws.cell(row=r, column=c)
        cl.value = label
        cl.font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        cl.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = thin_border

    def make_value_cell(ws, r, c, value, cpl):
        cv = ws.cell(row=r, column=c)
        cv.value = value
        cv.font = Font(name='微软雅黑', size=11, color=VALUE_C)
        cv.alignment = Alignment(wrap_text=True, vertical='top')
        cv.border = thin_border
        ws.row_dimensions[r].height = max(28, calc_lines(value, cpl) * 20)

    def pair_field(ws, r, label_col, label, val_col, value, cpl):
        """写一对 标签+值（不合并）"""
        make_label_cell(ws, r, label_col, label)
        make_value_cell(ws, r, val_col, value, cpl)

    def full_width_field(ws, r, label, value):
        """标签在A，值合并B:D"""
        make_label_cell(ws, r, 1, label)
        ws.merge_cells(f'B{r}:D{r}')
        make_value_cell(ws, r, 2, value, 140)  # B52 + C18 + D67 ≈ 137
        for cc in (3, 4):
            ws.cell(row=r, column=cc).border = thin_border

    def section_header(ws, r, title):
        ws.merge_cells(f'A{r}:D{r}')
        h = ws.cell(row=r, column=1)
        h.value = f"  {title}"
        h.font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        h.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
        h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        h.border = Border(
            left=Side(style='medium', color=DARK), right=Side(style='medium', color=DARK),
            top=Side(style='medium', color=DARK), bottom=Side(style='thin', color=DARK),
        )
        ws.row_dimensions[r].height = 28

    # 删除旧 sheet
    for s in list(wb.sheetnames):
        del wb[s]

    created = []
    seen_names = {}

    for values in rows_data:
        chezu = fmt_val(values[2])
        jixieshi = fmt_val(values[4])
        checi = fmt_val(values[3])

        raw = f"{jixieshi}{checi}"
        safe = re.sub(r'[\\/?*\[\]]', '·', raw)[:31]
        name = safe
        if name in seen_names:
            seen_names[name] += 1
            name = f"{safe[:28]}_{seen_names[name]}"
        else:
            seen_names[name] = 1

        ws_new = wb.create_sheet(title=name)

        # 标题
        ws_new.merge_cells('A1:D1')
        ws_new.row_dimensions[1].height = 48
        c = ws_new['A1']
        c.value = f"🚄 提示卡    {chezu}（{jixieshi}）"
        c.font = Font(name='微软雅黑', bold=True, size=20, color=DARK)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.fill = PatternFill(start_color=BG_TITLE, end_color=BG_TITLE, fill_type='solid')

        ws_new.row_dimensions[2].height = 6

        # === 分组1：基本运行信息（双列）===
        r = 3
        section_header(ws_new, r, "🚄 基本运行信息")
        r += 1

        pair_field(ws_new, r, 1, "日期", 2, fmt_val(values[1]), 50)
        pair_field(ws_new, r, 3, "车组号", 4, chezu, 65)
        r += 1

        pair_field(ws_new, r, 1, "机械师", 2, jixieshi, 50)
        pair_field(ws_new, r, 3, "所属配属", 4, fmt_val(values[14]), 65)
        r += 1

        full_width_field(ws_new, r, "交路名称", fmt_val(values[0]))
        r += 1
        full_width_field(ws_new, r, "动调电话", fmt_val(values[5]))
        r += 1

        ws_new.row_dimensions[r].height = 6
        r += 1

        # === 分组2：行车条件 ===
        section_header(ws_new, r, "🌤 行车条件")
        r += 1
        full_width_field(ws_new, r, "沿途天气", fmt_val(values[6])); r += 1
        full_width_field(ws_new, r, "行车限制条件重点提示", fmt_val(values[7])); r += 1
        full_width_field(ws_new, r, "交路运行情况", fmt_val(values[8])); r += 1

        ws_new.row_dimensions[r].height = 6
        r += 1

        # === 分组3：检修改造 ===
        section_header(ws_new, r, "🔧 检修改造")
        r += 1
        full_width_field(ws_new, r, "动车组源头质量和加装改造", fmt_val(values[9])); r += 1
        full_width_field(ws_new, r, "重点跟踪故障和关键配件更换", fmt_val(values[10])); r += 1
        full_width_field(ws_new, r, "检修信息", fmt_val(values[11])); r += 1

        ws_new.row_dimensions[r].height = 6
        r += 1

        # === 分组4：警示提示 ===
        section_header(ws_new, r, "⚠ 警示提示")
        r += 1
        full_width_field(ws_new, r, "典型案例警示", fmt_val(values[12])); r += 1
        full_width_field(ws_new, r, "作业风险提示", fmt_val(values[13])); r += 1

        # 列宽（跟用户调整一致）
        ws_new.column_dimensions['A'].width = 24
        ws_new.column_dimensions['B'].width = 52
        ws_new.column_dimensions['C'].width = 18
        ws_new.column_dimensions['D'].width = 67

        # 隐藏网格线 → 卡片效果
        ws_new.sheet_view.showGridLines = False

        created.append((name, chezu, jixieshi, checi))

    # === 查询页 ===
    ws_q = wb.create_sheet(title="查询页", index=0)
    ws_q.sheet_view.showGridLines = False

    ws_q.merge_cells('A1:D1')
    t = ws_q['A1']
    t.value = "提示卡查询 — 点击跳转"
    t.font = Font(name='微软雅黑', bold=True, size=16, color=DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_q.row_dimensions[1].height = 40

    hdr_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_align = Alignment(horizontal='center', vertical='center')

    for ci, txt in enumerate(["序号", "车组号", "姓名+车次", "跳转"], 1):
        c = ws_q.cell(row=2, column=ci)
        c.value = txt
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    link_font = Font(name='微软雅黑', size=11, color='0563C1', underline='single')
    normal_font = Font(name='微软雅黑', size=11)

    for i, (sn, cz, js, cc) in enumerate(created, 1):
        rn = i + 2
        ws_q.cell(row=rn, column=1).value = i
        ws_q.cell(row=rn, column=1).font = normal_font
        ws_q.cell(row=rn, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_q.cell(row=rn, column=1).border = thin_border

        ws_q.cell(row=rn, column=2).value = cz
        ws_q.cell(row=rn, column=2).font = normal_font
        ws_q.cell(row=rn, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws_q.cell(row=rn, column=2).border = thin_border

        ws_q.cell(row=rn, column=3).value = f"{js}{cc}"
        ws_q.cell(row=rn, column=3).font = normal_font
        ws_q.cell(row=rn, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws_q.cell(row=rn, column=3).border = thin_border

        ws_q.cell(row=rn, column=4).value = f'=HYPERLINK("#''{sn}''!A1","点击查看")'
        ws_q.cell(row=rn, column=4).font = link_font
        ws_q.cell(row=rn, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws_q.cell(row=rn, column=4).border = thin_border
        ws_q.row_dimensions[rn].height = 24

    ws_q.column_dimensions['A'].width = 8
    ws_q.column_dimensions['B'].width = 12
    ws_q.column_dimensions['C'].width = 22
    ws_q.column_dimensions['D'].width = 14

    base, ext = os.path.splitext(input_path)
    out_path = f"{base}（分表）{ext}"
    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) >= 2:
        inputs = sys.argv[1:]
    else:
        print("=" * 50)
        print("  提示卡分表工具 v2.0")
        print("=" * 50)
        inp = input("\n请拖入或输入 Excel 文件路径：\n> ").strip().strip('"')
        if not inp:
            print("未输入文件，退出。")
            input("\n按回车键退出...")
            return
        inputs = [inp]

    for fp in inputs:
        path = fp.strip().strip('"')
        if not os.path.isfile(path):
            print(f"  ✗ 跳过（文件不存在）：{path}")
            continue
        try:
            out = split_prompt_card(path)
            print(f"\n  ✓ 生成成功：{out}")
            print(f"    文件大小：{os.path.getsize(out):,} 字节")
        except Exception as e:
            print(f"\n  ✗ 处理失败：{path}")
            print(f"    错误：{e}")

    input("\n完成！按回车键退出...")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        input("\n发生未预期错误，按回车键退出...")
